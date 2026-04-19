"""
Build ClearLaunch Step 6 Projections Template (.xlsx) — 4-tab workbook
with live formulas, yellow input cells, green calculated cells, and cross-sheet refs.
Industry-agnostic version.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ========================================
# STYLE DEFINITIONS
# ========================================

# Fills
YELLOW_FILL = PatternFill(start_color="FFFFFDE7", end_color="FFFFFDE7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFE8F5EF", end_color="FFE8F5EF", fill_type="solid")
DARK_FILL = PatternFill(start_color="FF121718", end_color="FF121718", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
LGRAY_FILL = PatternFill(start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid")
BRAND_GREEN_FILL = PatternFill(start_color="FF1B9B5E", end_color="FF1B9B5E", fill_type="solid")
BRAND_ORANGE_FILL = PatternFill(start_color="FFD4880F", end_color="FFD4880F", fill_type="solid")
LGGREEN_FILL = PatternFill(start_color="FFE5F5EC", end_color="FFE5F5EC", fill_type="solid")

# Fonts
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="121718")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1B9B5E")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", size=10, bold=False, color="121718")
LABEL_BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="121718")
INPUT_FONT = Font(name="Calibri", size=10, bold=True, color="0000FF")
CALC_FONT = Font(name="Calibri", size=10, bold=False, color="1B9B5E")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
WHITE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BADGE_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
LINK_FONT = Font(name="Calibri", size=10, bold=False, color="1B9B5E")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="A8D9BD"),
    right=Side(style="thin", color="A8D9BD"),
    top=Side(style="thin", color="A8D9BD"),
    bottom=Side(style="thin", color="A8D9BD"),
)

# Alignment
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# Number formats
CURRENCY = '"$"#,##0'
CURRENCY_DEC = '"$"#,##0.00'
PERCENT = '0.0%'
NUMBER = '#,##0'
NUMBER_DEC = '#,##0.0'
ROAS = '0.0"x"'


def set_header_row(ws, row, headers, col_start=1):
    """Apply dark header styling to a row."""
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + j, value=h)
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def set_label(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = LABEL_BOLD_FONT if bold else LABEL_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    return cell


def set_input(ws, row, col, value=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = YELLOW_FILL
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def set_calc(ws, row, col, formula, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = CALC_FONT
    cell.fill = GREEN_FILL
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def set_linked(ws, row, col, formula, fmt=None):
    """For Summary tab cross-sheet references."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = LINK_FONT
    cell.fill = GREEN_FILL
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def set_note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    return cell


def alt_row_fill(row_idx):
    """Alternating row fill: white (even data rows) / light gray (odd)."""
    return LGRAY_FILL if row_idx % 2 == 0 else WHITE_FILL


def apply_row_fill(ws, row, cols, fill):
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).fill = fill


# ========================================
# BUILD WORKBOOK
# ========================================
wb = Workbook()

# ========================================
# TAB 1: SUMMARY
# ========================================
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.sheet_properties.tabColor = "1B9B5E"

# Column widths
ws_sum.column_dimensions["A"].width = 28
ws_sum.column_dimensions["B"].width = 20
ws_sum.column_dimensions["C"].width = 20
ws_sum.column_dimensions["D"].width = 20
ws_sum.column_dimensions["E"].width = 25

# Title
ws_sum.cell(row=1, column=1, value="[CLIENT NAME] — CHANNEL STRATEGY SUMMARY").font = TITLE_FONT
ws_sum.merge_cells("A1:E1")
ws_sum.cell(row=2, column=1,
            value="Green values = linked from detail tabs. Yellow cells = adjustable inputs.").font = NOTE_FONT
ws_sum.merge_cells("A2:E2")

# ---- CHANNEL ROLES ----
ws_sum.cell(row=4, column=1, value="CHANNEL ROLES").font = SECTION_FONT

ws_sum.cell(row=5, column=1, value="[Primary Channel 1]").font = LABEL_BOLD_FONT
ws_sum.cell(row=5, column=2, value="[Role Title]").font = Font(name="Calibri", size=10, italic=True, color="1B9B5E")
ws_sum.cell(row=6, column=1, value="[Role description — 1-2 sentences]").font = NOTE_FONT
ws_sum.merge_cells("A6:E6")

ws_sum.cell(row=7, column=1, value="[Primary Channel 2]").font = LABEL_BOLD_FONT
ws_sum.cell(row=7, column=2, value="[Role Title]").font = Font(name="Calibri", size=10, italic=True, color="1B9B5E")
ws_sum.cell(row=8, column=1, value="[Role description — 1-2 sentences]").font = NOTE_FONT
ws_sum.merge_cells("A8:E8")

ws_sum.cell(row=9, column=1, value="[Supporting Channel]").font = LABEL_BOLD_FONT
ws_sum.cell(row=9, column=2, value="[Role Title]").font = Font(name="Calibri", size=10, italic=True, color="1B9B5E")
ws_sum.cell(row=10, column=1, value="[Role description — 1-2 sentences]").font = NOTE_FONT
ws_sum.merge_cells("A10:E10")

# ---- INVESTMENT OVERVIEW ----
ws_sum.cell(row=12, column=1, value="INVESTMENT OVERVIEW").font = SECTION_FONT

set_header_row(ws_sum, 13, ["Channel", "Monthly Retainer", "Monthly Ad Spend", "Total Monthly", "Notes"])

inv_rows = [
    (14, "[Channel 1]", "[retainer]", "[ad spend]", "[total]", "[notes]"),
    (15, "[Channel 2]", "[retainer]", "[ad spend]", "[total]", "[notes]"),
    (16, "[Channel 3]", "[retainer]", "[ad spend]", "[total]", "[notes]"),
]
for row, ch, ret, ad, tot, notes in inv_rows:
    set_label(ws_sum, row, 1, ch)
    set_input(ws_sum, row, 2, ret)
    set_input(ws_sum, row, 3, ad)
    set_input(ws_sum, row, 4, tot)
    set_note(ws_sum, row, 5, notes)

# Totals row
set_label(ws_sum, 17, 1, "TOTAL", bold=True)
ws_sum.cell(row=17, column=1).fill = LGGREEN_FILL
set_calc(ws_sum, 17, 2, "=SUM(B14:B16)", CURRENCY)
ws_sum.cell(row=17, column=2).fill = LGGREEN_FILL
set_calc(ws_sum, 17, 3, "=SUM(C14:C16)", CURRENCY)
ws_sum.cell(row=17, column=3).fill = LGGREEN_FILL
set_calc(ws_sum, 17, 4, "=SUM(D14:D16)", CURRENCY)
ws_sum.cell(row=17, column=4).fill = LGGREEN_FILL
ws_sum.cell(row=17, column=4).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

# ---- GOOGLE ADS PROJECTED PERFORMANCE ----
ws_sum.cell(row=19, column=1, value="GOOGLE ADS PROJECTED PERFORMANCE").font = SECTION_FONT

# Badge
ws_sum.cell(row=19, column=3, value="DATA-BACKED ESTIMATE").font = BADGE_FONT
ws_sum.cell(row=19, column=3).fill = BRAND_GREEN_FILL

set_header_row(ws_sum, 20, ["Metric", "[Service Line 1]", "[Service Line 2]"])

google_metrics = [
    (21, "Monthly Ad Budget"),
    (22, "Avg CPC"),
    (23, "Monthly Clicks"),
    (24, "LP Conversion Rate"),
    (25, "Monthly Leads"),
    (26, "Sales Closing Rate"),
    (27, "New Clients/Month"),
    (28, "Monthly Revenue"),
    (29, "Ad Spend ROAS"),
    (30, "Cost Per Acquisition"),
]

for row, metric in google_metrics:
    set_label(ws_sum, row, 1, metric)
    # Cross-sheet references to Google Ads tab
    # Map: Summary rows -> Google Ads rows
    ga_map = {
        21: ("B9", "C9", CURRENCY),       # Ad Budget
        22: ("B8", "C8", CURRENCY_DEC),    # CPC
        23: ("B20", "C20", NUMBER),        # Clicks
        24: ("B10", "C10", PERCENT),       # Conv Rate
        25: ("B21", "C21", NUMBER_DEC),    # Leads
        26: ("B11", "C11", PERCENT),       # Close Rate
        27: ("B23", "C23", NUMBER_DEC),    # Clients
        28: ("B27", "C27", CURRENCY),      # Revenue
        29: ("B28", "C28", ROAS),          # ROAS
        30: ("B24", "C24", CURRENCY),      # CPA
    }
    b_ref, c_ref, fmt = ga_map[row]
    set_linked(ws_sum, row, 2, f"='Google Ads'!{b_ref}", fmt)
    set_linked(ws_sum, row, 3, f"='Google Ads'!{c_ref}", fmt)

# ---- META ADS PROJECTED PERFORMANCE ----
ws_sum.cell(row=32, column=1, value="META ADS PROJECTED PERFORMANCE").font = SECTION_FONT

# Badge
ws_sum.cell(row=32, column=3, value="MODELED ESTIMATE").font = BADGE_FONT
ws_sum.cell(row=32, column=3).fill = BRAND_ORANGE_FILL

set_header_row(ws_sum, 33, ["Metric", "[Service Line 1]", "[Service Line 2]"])

meta_metrics = [
    (34, "Monthly Ad Budget"),
    (35, "CPM"),
    (36, "Effective CPC"),
    (37, "Monthly Leads"),
    (38, "New Clients/Month"),
    (39, "Monthly Revenue"),
    (40, "Ad Spend ROAS"),
    (41, "Cost Per Acquisition"),
]

for row, metric in meta_metrics:
    set_label(ws_sum, row, 1, metric)
    ma_map = {
        34: ("B9", "C9", CURRENCY),        # Monthly Budget
        35: ("B10", "C10", CURRENCY_DEC),   # CPM
        36: ("B12", "C12", CURRENCY_DEC),   # Effective CPC
        37: ("B27", "C27", NUMBER_DEC),     # Leads
        38: ("B29", "C29", NUMBER_DEC),     # Clients
        39: ("B31", "C31", CURRENCY),       # Revenue
        40: ("B32", "C32", ROAS),           # ROAS
        41: ("B30", "C30", CURRENCY),       # CPA
    }
    b_ref, c_ref, fmt = ma_map[row]
    set_linked(ws_sum, row, 2, f"='Meta Ads'!{b_ref}", fmt)
    set_linked(ws_sum, row, 3, f"='Meta Ads'!{c_ref}", fmt)

# ---- KEY TAKEAWAYS ----
ws_sum.cell(row=43, column=1, value="KEY TAKEAWAYS").font = SECTION_FONT

takeaways = [
    "[Takeaway 1: Primary channel role and foundation]",
    "[Takeaway 2: Secondary channel complement]",
    "[Takeaway 3: Supporting channel for demand generation]",
    "[Takeaway 4: Funnel architecture across channels]",
    "[Takeaway 5: Budget control and scaling strategy]",
]
for i, t in enumerate(takeaways):
    ws_sum.cell(row=44 + i, column=1, value=f"{i+1}. {t}").font = NOTE_FONT
    ws_sum.merge_cells(f"A{44+i}:E{44+i}")


# ========================================
# TAB 2: GOOGLE ADS (Data-Backed)
# ========================================
ws_ga = wb.create_sheet("Google Ads")
ws_ga.sheet_properties.tabColor = "1B9B5E"

ws_ga.column_dimensions["A"].width = 28
ws_ga.column_dimensions["B"].width = 20
ws_ga.column_dimensions["C"].width = 20
ws_ga.column_dimensions["D"].width = 25

# Title
ws_ga.cell(row=1, column=1, value="[CLIENT NAME] — GOOGLE ADS PROJECTIONS").font = TITLE_FONT
ws_ga.merge_cells("A1:D1")
ws_ga.cell(row=2, column=1, value="DATA-BACKED ESTIMATE — Built from real keyword data (Step 3 Market Research)").font = BADGE_FONT
ws_ga.cell(row=2, column=1).fill = BRAND_GREEN_FILL
ws_ga.merge_cells("A2:D2")
ws_ga.cell(row=3, column=1,
           value="Yellow = adjustable inputs (blue text). Green = auto-calculated.").font = NOTE_FONT

# Header
set_header_row(ws_ga, 5, ["Metric", "[Service Line 1]", "[Service Line 2]", "Notes"])

# ---- INPUTS ----
set_label(ws_ga, 6, 1, "DFY Management Retainer", bold=True)
set_input(ws_ga, 6, 2, 0, CURRENCY)
set_note(ws_ga, 6, 4, "Fixed monthly management fee")

set_label(ws_ga, 7, 1, "")  # spacer

set_label(ws_ga, 8, 1, "Avg CPC")
set_input(ws_ga, 8, 2, 0, CURRENCY_DEC)
set_input(ws_ga, 8, 3, 0, CURRENCY_DEC)
set_note(ws_ga, 8, 4, "From Step 3 keyword data")

set_label(ws_ga, 9, 1, "Monthly Ad Budget")
set_input(ws_ga, 9, 2, 0, CURRENCY)
set_input(ws_ga, 9, 3, 0, CURRENCY)
set_note(ws_ga, 9, 4, "Adjustable input")

set_label(ws_ga, 10, 1, "LP Conversion Rate")
set_input(ws_ga, 10, 2, 0, PERCENT)
set_input(ws_ga, 10, 3, 0, PERCENT)
set_note(ws_ga, 10, 4, "Landing page visitor → lead")

set_label(ws_ga, 11, 1, "Sales Closing Rate")
set_input(ws_ga, 11, 2, 0, PERCENT)
set_input(ws_ga, 11, 3, 0, PERCENT)
set_note(ws_ga, 11, 4, "Lead → paying client")

set_label(ws_ga, 12, 1, "")  # spacer

set_label(ws_ga, 13, 1, "REVENUE PER CLIENT", bold=True)
set_label(ws_ga, 14, 1, "[Service Line 1] Price")
set_input(ws_ga, 14, 2, 0, CURRENCY)
set_note(ws_ga, 14, 4, "From Step 5 Offer Dev")

set_label(ws_ga, 15, 1, "[Service Line 2] Rate")
set_input(ws_ga, 15, 3, 0, CURRENCY)
set_note(ws_ga, 15, 4, "Hourly/per-unit rate")

set_label(ws_ga, 16, 1, "Est. Units/Engagement")
set_input(ws_ga, 16, 3, 0, NUMBER)
set_note(ws_ga, 16, 4, "Hours, sessions, units per engagement")

set_label(ws_ga, 17, 1, "Revenue Per Client")
set_calc(ws_ga, 17, 2, "=B14", CURRENCY)
set_calc(ws_ga, 17, 3, "=C15*C16", CURRENCY)
set_note(ws_ga, 17, 4, "Line 1 = flat price; Line 2 = rate × units")

# ---- SPACER ----
set_label(ws_ga, 18, 1, "")

# ---- CALCULATED METRICS ----
set_label(ws_ga, 19, 1, "PROJECTED PERFORMANCE", bold=True)

set_label(ws_ga, 20, 1, "Monthly Clicks")
set_calc(ws_ga, 20, 2, "=IF(B8>0,B9/B8,0)", NUMBER)
set_calc(ws_ga, 20, 3, "=IF(C8>0,C9/C8,0)", NUMBER)
set_note(ws_ga, 20, 4, "= Budget ÷ CPC")

set_label(ws_ga, 21, 1, "Monthly Leads")
set_calc(ws_ga, 21, 2, "=B20*B10", NUMBER_DEC)
set_calc(ws_ga, 21, 3, "=C20*C10", NUMBER_DEC)
set_note(ws_ga, 21, 4, "= Clicks × Conv Rate")

set_label(ws_ga, 22, 1, "Cost Per Lead")
set_calc(ws_ga, 22, 2, "=IF(B21>0,B9/B21,0)", CURRENCY_DEC)
set_calc(ws_ga, 22, 3, "=IF(C21>0,C9/C21,0)", CURRENCY_DEC)
set_note(ws_ga, 22, 4, "= Budget ÷ Leads")

set_label(ws_ga, 23, 1, "New Clients/Month")
set_calc(ws_ga, 23, 2, "=B21*B11", NUMBER_DEC)
set_calc(ws_ga, 23, 3, "=C21*C11", NUMBER_DEC)
set_note(ws_ga, 23, 4, "= Leads × Close Rate")

set_label(ws_ga, 24, 1, "Cost Per Acquisition")
set_calc(ws_ga, 24, 2, "=IF(B23>0,B9/B23,0)", CURRENCY_DEC)
set_calc(ws_ga, 24, 3, "=IF(C23>0,C9/C23,0)", CURRENCY_DEC)
set_note(ws_ga, 24, 4, "= Ad Spend ÷ Clients")

set_label(ws_ga, 25, 1, "")  # spacer

set_label(ws_ga, 26, 1, "REVENUE & PROFITABILITY", bold=True)

set_label(ws_ga, 27, 1, "Monthly Revenue")
set_calc(ws_ga, 27, 2, "=B23*B17", CURRENCY)
set_calc(ws_ga, 27, 3, "=C23*C17", CURRENCY)
set_note(ws_ga, 27, 4, "= Clients × Rev/Client")

set_label(ws_ga, 28, 1, "Ad Spend ROAS")
set_calc(ws_ga, 28, 2, "=IF(B9>0,B27/B9,0)", ROAS)
set_calc(ws_ga, 28, 3, "=IF(C9>0,C27/C9,0)", ROAS)
set_note(ws_ga, 28, 4, "= Revenue ÷ Ad Spend only")

set_label(ws_ga, 29, 1, "Monthly Profit (ad spend only)")
set_calc(ws_ga, 29, 2, "=B27-B9", CURRENCY)
set_calc(ws_ga, 29, 3, "=C27-C9", CURRENCY)
set_note(ws_ga, 29, 4, "= Revenue − Ad Spend")

set_label(ws_ga, 30, 1, "Profit Margin (ad spend only)")
set_calc(ws_ga, 30, 2, "=IF(B27>0,B29/B27,0)", PERCENT)
set_calc(ws_ga, 30, 3, "=IF(C27>0,C29/C27,0)", PERCENT)

# ---- COMBINED PROFITABILITY (includes retainer) ----
set_label(ws_ga, 32, 1, "")
set_label(ws_ga, 33, 1, "COMBINED CHANNEL PROFITABILITY", bold=True)
set_header_row(ws_ga, 34, ["Metric", "", "", "Combined"])

set_label(ws_ga, 35, 1, "Total Ad Spend")
set_calc(ws_ga, 35, 4, "=B9+C9", CURRENCY)

set_label(ws_ga, 36, 1, "Management Retainer")
set_calc(ws_ga, 36, 4, "=B6", CURRENCY)

set_label(ws_ga, 37, 1, "Total Investment")
set_calc(ws_ga, 37, 4, "=D35+D36", CURRENCY)
ws_ga.cell(row=37, column=4).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

set_label(ws_ga, 38, 1, "Total Revenue")
set_calc(ws_ga, 38, 4, "=B27+C27", CURRENCY)

set_label(ws_ga, 39, 1, "Total New Clients")
set_calc(ws_ga, 39, 4, "=B23+C23", NUMBER_DEC)

set_label(ws_ga, 40, 1, "CPA (total investment)")
set_calc(ws_ga, 40, 4, "=IF(D39>0,D37/D39,0)", CURRENCY_DEC)

set_label(ws_ga, 41, 1, "True ROAS (incl. retainer)")
set_calc(ws_ga, 41, 4, "=IF(D37>0,D38/D37,0)", ROAS)
ws_ga.cell(row=41, column=4).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

set_label(ws_ga, 42, 1, "Profit (after all costs)")
set_calc(ws_ga, 42, 4, "=D38-D37", CURRENCY)

set_label(ws_ga, 43, 1, "Profit Margin")
set_calc(ws_ga, 43, 4, "=IF(D38>0,D42/D38,0)", PERCENT)


# ========================================
# TAB 3: META ADS (Modeled)
# ========================================
ws_ma = wb.create_sheet("Meta Ads")
ws_ma.sheet_properties.tabColor = "D4880F"

ws_ma.column_dimensions["A"].width = 28
ws_ma.column_dimensions["B"].width = 20
ws_ma.column_dimensions["C"].width = 20
ws_ma.column_dimensions["D"].width = 25

# Title
ws_ma.cell(row=1, column=1, value="[CLIENT NAME] — META ADS PROJECTIONS").font = TITLE_FONT
ws_ma.merge_cells("A1:D1")
ws_ma.cell(row=2, column=1, value="MODELED ESTIMATE — Built from industry CPM/CTR benchmarks").font = BADGE_FONT
ws_ma.cell(row=2, column=1).fill = BRAND_ORANGE_FILL
ws_ma.merge_cells("A2:D2")
ws_ma.cell(row=3, column=1,
           value="Yellow = adjustable inputs (blue text). Green = auto-calculated.").font = NOTE_FONT

# Header
set_header_row(ws_ma, 5, ["Metric", "[Service Line 1]", "[Service Line 2]", "Notes"])

# ---- INPUTS ----
set_label(ws_ma, 6, 1, "DFY Management Retainer", bold=True)
set_input(ws_ma, 6, 2, 0, CURRENCY)
set_note(ws_ma, 6, 4, "Fixed monthly management fee")

set_label(ws_ma, 7, 1, "")  # spacer

set_label(ws_ma, 8, 1, "Daily Ad Budget")
set_input(ws_ma, 8, 2, 0, CURRENCY)
set_input(ws_ma, 8, 3, 0, CURRENCY)
set_note(ws_ma, 8, 4, "Per-day ad spend")

set_label(ws_ma, 9, 1, "Monthly Ad Budget")
set_calc(ws_ma, 9, 2, "=B8*30", CURRENCY)
set_calc(ws_ma, 9, 3, "=C8*30", CURRENCY)
set_note(ws_ma, 9, 4, "= Daily × 30")

set_label(ws_ma, 10, 1, "CPM")
set_input(ws_ma, 10, 2, 0, CURRENCY_DEC)
set_input(ws_ma, 10, 3, 0, CURRENCY_DEC)
set_note(ws_ma, 10, 4, "Cost per 1,000 impressions (benchmark)")

set_label(ws_ma, 11, 1, "CTR")
set_input(ws_ma, 11, 2, 0, PERCENT)
set_input(ws_ma, 11, 3, 0, PERCENT)
set_note(ws_ma, 11, 4, "Click-through rate (benchmark)")

set_label(ws_ma, 12, 1, "Effective CPC")
set_calc(ws_ma, 12, 2, "=IF(B11>0,B10/(B11*1000),0)", CURRENCY_DEC)
set_calc(ws_ma, 12, 3, "=IF(C11>0,C10/(C11*1000),0)", CURRENCY_DEC)
set_note(ws_ma, 12, 4, "= CPM ÷ (CTR × 1000)")

set_label(ws_ma, 13, 1, "")  # spacer

set_label(ws_ma, 14, 1, "CONVERSION INPUTS", bold=True)

set_label(ws_ma, 15, 1, "LP Conversion Rate")
set_input(ws_ma, 15, 2, 0, PERCENT)
set_input(ws_ma, 15, 3, 0, PERCENT)
set_note(ws_ma, 15, 4, "Landing page visitor → lead")

set_label(ws_ma, 16, 1, "Sales Closing Rate")
set_input(ws_ma, 16, 2, 0, PERCENT)
set_input(ws_ma, 16, 3, 0, PERCENT)
set_note(ws_ma, 16, 4, "Lead → paying client")

set_label(ws_ma, 17, 1, "")  # spacer

set_label(ws_ma, 18, 1, "REVENUE", bold=True)

set_label(ws_ma, 19, 1, "Revenue Per Client")
set_input(ws_ma, 19, 2, 0, CURRENCY)
set_input(ws_ma, 19, 3, 0, CURRENCY)
set_note(ws_ma, 19, 4, "From Step 5 Offer Dev")

set_label(ws_ma, 20, 1, "")  # spacer

# ---- CALCULATED METRICS ----
set_label(ws_ma, 21, 1, "PROJECTED PERFORMANCE", bold=True)

set_label(ws_ma, 22, 1, "Monthly Impressions")
set_calc(ws_ma, 22, 2, "=IF(B10>0,B9/B10*1000,0)", NUMBER)
set_calc(ws_ma, 22, 3, "=IF(C10>0,C9/C10*1000,0)", NUMBER)
set_note(ws_ma, 22, 4, "= Budget ÷ CPM × 1000")

set_label(ws_ma, 23, 1, "Monthly Clicks")
set_calc(ws_ma, 23, 2, "=B22*B11", NUMBER)
set_calc(ws_ma, 23, 3, "=C22*C11", NUMBER)
set_note(ws_ma, 23, 4, "= Impressions × CTR")

set_label(ws_ma, 24, 1, "Cost Per Click")
set_calc(ws_ma, 24, 2, "=IF(B23>0,B9/B23,0)", CURRENCY_DEC)
set_calc(ws_ma, 24, 3, "=IF(C23>0,C9/C23,0)", CURRENCY_DEC)

set_label(ws_ma, 25, 1, "")  # spacer

set_label(ws_ma, 26, 1, "CONVERSION FUNNEL", bold=True)

set_label(ws_ma, 27, 1, "Monthly Leads")
set_calc(ws_ma, 27, 2, "=B23*B15", NUMBER_DEC)
set_calc(ws_ma, 27, 3, "=C23*C15", NUMBER_DEC)
set_note(ws_ma, 27, 4, "= Clicks × LP Conv Rate")

set_label(ws_ma, 28, 1, "Cost Per Lead")
set_calc(ws_ma, 28, 2, "=IF(B27>0,B9/B27,0)", CURRENCY_DEC)
set_calc(ws_ma, 28, 3, "=IF(C27>0,C9/C27,0)", CURRENCY_DEC)

set_label(ws_ma, 29, 1, "New Clients/Month")
set_calc(ws_ma, 29, 2, "=B27*B16", NUMBER_DEC)
set_calc(ws_ma, 29, 3, "=C27*C16", NUMBER_DEC)
set_note(ws_ma, 29, 4, "= Leads × Close Rate")

set_label(ws_ma, 30, 1, "Cost Per Acquisition")
set_calc(ws_ma, 30, 2, "=IF(B29>0,B9/B29,0)", CURRENCY_DEC)
set_calc(ws_ma, 30, 3, "=IF(C29>0,C9/C29,0)", CURRENCY_DEC)
set_note(ws_ma, 30, 4, "= Ad Spend ÷ Clients")

set_label(ws_ma, 31, 1, "Monthly Revenue")
set_calc(ws_ma, 31, 2, "=B29*B19", CURRENCY)
set_calc(ws_ma, 31, 3, "=C29*C19", CURRENCY)
set_note(ws_ma, 31, 4, "= Clients × Rev/Client")

set_label(ws_ma, 32, 1, "Ad Spend ROAS")
set_calc(ws_ma, 32, 2, "=IF(B9>0,B31/B9,0)", ROAS)
set_calc(ws_ma, 32, 3, "=IF(C9>0,C31/C9,0)", ROAS)
set_note(ws_ma, 32, 4, "= Revenue ÷ Ad Spend only")

set_label(ws_ma, 33, 1, "Monthly Profit (ad spend only)")
set_calc(ws_ma, 33, 2, "=B31-B9", CURRENCY)
set_calc(ws_ma, 33, 3, "=C31-C9", CURRENCY)

set_label(ws_ma, 34, 1, "Profit Margin (ad spend only)")
set_calc(ws_ma, 34, 2, "=IF(B31>0,B33/B31,0)", PERCENT)
set_calc(ws_ma, 34, 3, "=IF(C31>0,C33/C31,0)", PERCENT)

# ---- COMBINED PROFITABILITY (includes retainer) — MATCHING GOOGLE ADS PATTERN ----
set_label(ws_ma, 36, 1, "")
set_label(ws_ma, 37, 1, "COMBINED CHANNEL PROFITABILITY", bold=True)
set_header_row(ws_ma, 38, ["Metric", "", "", "Combined"])

set_label(ws_ma, 39, 1, "Total Ad Spend")
set_calc(ws_ma, 39, 4, "=B9+C9", CURRENCY)

set_label(ws_ma, 40, 1, "Management Retainer")
set_calc(ws_ma, 40, 4, "=B6", CURRENCY)

set_label(ws_ma, 41, 1, "Total Investment")
set_calc(ws_ma, 41, 4, "=D39+D40", CURRENCY)
ws_ma.cell(row=41, column=4).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

set_label(ws_ma, 42, 1, "Total Revenue")
set_calc(ws_ma, 42, 4, "=B31+C31", CURRENCY)

set_label(ws_ma, 43, 1, "Total New Clients")
set_calc(ws_ma, 43, 4, "=B29+C29", NUMBER_DEC)

set_label(ws_ma, 44, 1, "CPA (total investment)")
set_calc(ws_ma, 44, 4, "=IF(D43>0,D41/D43,0)", CURRENCY_DEC)

set_label(ws_ma, 45, 1, "True ROAS (incl. retainer)")
set_calc(ws_ma, 45, 4, "=IF(D41>0,D42/D41,0)", ROAS)
ws_ma.cell(row=45, column=4).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

set_label(ws_ma, 46, 1, "Profit (after all costs)")
set_calc(ws_ma, 46, 4, "=D42-D41", CURRENCY)

set_label(ws_ma, 47, 1, "Profit Margin")
set_calc(ws_ma, 47, 4, "=IF(D42>0,D46/D42,0)", PERCENT)

# ---- BENCHMARK SOURCES ----
ws_ma.cell(row=49, column=1, value="BENCHMARK SOURCES").font = SECTION_FONT
set_header_row(ws_ma, 50, ["Source", "Metric", "Value", "Notes"])

benchmarks = [
    ("[Source 1]", "CPM", "[range]", "[industry/context]"),
    ("[Source 2]", "CTR", "[range]", "[industry/context]"),
    ("[Source 3]", "LP Conv Rate", "[range]", "[industry/context]"),
    ("[Source 4]", "Close Rate", "[range]", "[industry/context]"),
]
for i, (src, metric, val, notes) in enumerate(benchmarks):
    row = 51 + i
    set_label(ws_ma, row, 1, src)
    set_label(ws_ma, row, 2, metric)
    set_label(ws_ma, row, 3, val)
    set_note(ws_ma, row, 4, notes)


# ========================================
# TAB 4: CONTENT & SEO
# ========================================
ws_cs = wb.create_sheet("Content & SEO")
ws_cs.sheet_properties.tabColor = "1B9B5E"

ws_cs.column_dimensions["A"].width = 30
ws_cs.column_dimensions["B"].width = 12
ws_cs.column_dimensions["C"].width = 15
ws_cs.column_dimensions["D"].width = 18
ws_cs.column_dimensions["E"].width = 25

# Title
ws_cs.cell(row=1, column=1, value="[CLIENT NAME] — CONTENT & SEO FOUNDATION").font = TITLE_FONT
ws_cs.merge_cells("A1:E1")
ws_cs.cell(row=2, column=1,
           value="Page inventory and SEO services to support channel strategy.").font = NOTE_FONT

# ---- CONTENT MARKETING ----
ws_cs.cell(row=4, column=1, value="CONTENT MARKETING").font = SECTION_FONT
set_header_row(ws_cs, 5, ["Deliverable", "Pages", "Cost/Page", "Subtotal", "Notes"])

categories = [
    "[Page Category 1]",
    "[Page Category 2]",
    "[Page Category 3]",
    "[Page Category 4]",
    "[Page Category 5]",
    "[Page Category 6]",
]
for i, cat in enumerate(categories):
    row = 6 + i
    set_label(ws_cs, row, 1, cat)
    set_input(ws_cs, row, 2, 0, NUMBER)
    set_input(ws_cs, row, 3, 0, CURRENCY)
    set_calc(ws_cs, row, 4, f"=B{row}*C{row}", CURRENCY)
    set_note(ws_cs, row, 5, "[description]")

# Subtotal
sub_row = 6 + len(categories)
set_label(ws_cs, sub_row, 1, "Content Marketing Subtotal", bold=True)
ws_cs.cell(row=sub_row, column=1).fill = LGGREEN_FILL
set_calc(ws_cs, sub_row, 2, f"=SUM(B6:B{sub_row-1})", NUMBER)
ws_cs.cell(row=sub_row, column=2).fill = LGGREEN_FILL
set_calc(ws_cs, sub_row, 4, f"=SUM(D6:D{sub_row-1})", CURRENCY)
ws_cs.cell(row=sub_row, column=4).fill = LGGREEN_FILL
ws_cs.cell(row=sub_row, column=4).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

# ---- SEO SERVICES ----
seo_start = sub_row + 2
ws_cs.cell(row=seo_start, column=1, value="SEO SERVICES").font = SECTION_FONT
set_header_row(ws_cs, seo_start + 1, ["Service", "Details", "Monthly Cost", "", ""])

seo_services = [
    ("[Service 1]", "[description]"),
    ("[Service 2]", "[description]"),
    ("[Service 3]", "[description]"),
]
for i, (svc, desc) in enumerate(seo_services):
    row = seo_start + 2 + i
    set_label(ws_cs, row, 1, svc)
    set_note(ws_cs, row, 2, desc)
    set_input(ws_cs, row, 3, 0, CURRENCY)

# ---- ONE-TIME COSTS ----
ot_start = seo_start + 2 + len(seo_services) + 1
ws_cs.cell(row=ot_start, column=1, value="ONE-TIME COSTS").font = SECTION_FONT

set_label(ws_cs, ot_start + 1, 1, "[One-time service]")
set_input(ws_cs, ot_start + 1, 3, 0, CURRENCY)
set_note(ws_cs, ot_start + 1, 4, "[description]")

# ---- INVESTMENT SUMMARY ----
inv_start = ot_start + 3
ws_cs.cell(row=inv_start, column=1, value="INVESTMENT SUMMARY").font = SECTION_FONT

set_label(ws_cs, inv_start + 1, 1, "Content Marketing (one-time)", bold=True)
set_calc(ws_cs, inv_start + 1, 3, f"=D{sub_row}", CURRENCY)

set_label(ws_cs, inv_start + 2, 1, "One-time Costs")
set_calc(ws_cs, inv_start + 2, 3, f"=C{ot_start + 1}", CURRENCY)

set_label(ws_cs, inv_start + 3, 1, "Total One-Time Investment", bold=True)
ws_cs.cell(row=inv_start + 3, column=1).fill = LGGREEN_FILL
set_calc(ws_cs, inv_start + 3, 3, f"=C{inv_start+1}+C{inv_start+2}", CURRENCY)
ws_cs.cell(row=inv_start + 3, column=3).fill = LGGREEN_FILL
ws_cs.cell(row=inv_start + 3, column=3).font = Font(name="Calibri", size=10, bold=True, color="1B9B5E")

set_label(ws_cs, inv_start + 5, 1, "Monthly SEO Services")
seo_range_end = seo_start + 1 + len(seo_services)
set_calc(ws_cs, inv_start + 5, 3, f"=SUM(C{seo_start+2}:C{seo_range_end})", CURRENCY)
set_note(ws_cs, inv_start + 5, 4, "Ongoing monthly cost")


# ========================================
# SAVE
# ========================================
wb.save("Frameworks/ClearLaunch_Step6_Projections_Template.xlsx")
print("Created: Frameworks/ClearLaunch_Step6_Projections_Template.xlsx (4 tabs)")
